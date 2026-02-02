# engine.py
from pathlib import Path
from openpyxl import load_workbook, Workbook


def run_calculation(input_path: Path, output_path: Path) -> None:

    #Load workbook and value of the cell A1
    wb_in = load_workbook(input_path, data_only=True)
    ws_in = wb_in.active
    value = ws_in["A1"].value

    #Check for errors
    if value is None:
        raise ValueError("Cell A1 is empty. Please enter a number in A1.")
    if not isinstance(value, (int, float)):
        raise ValueError(f"Cell A1 must be a number. Got: {value!r}")

    #Simple "Secret Calculation"
    doubled = value * 2

    #creates new workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Output"

    #adds the relevant data in relevant cells
    ws_out["A1"] = "Input A1"
    ws_out["B1"] = value
    ws_out["A2"] = "Doubled"
    ws_out["B2"] = doubled

    # saves the data to the file in outputs
    wb_out.save(output_path)
